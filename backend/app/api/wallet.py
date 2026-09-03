"""资金链路接口（阶段四 T4-1~T4-5 + 微信自动打款）。

资金安全要点，全部遵循《技术机制预研.md》：
- T4-1 佣金入账：佣金写 commissions + 余额加 balances.available_balance，同一事务内完成；
  commissions.assignment_id 唯一索引保证幂等（即使并发重复触发也只入账一次）。
- T4-2 提现申请：校验 amount <= available_balance 且已绑卡，事务内「扣余额 + 写提现记录」，免二次审核；
  提交后同步调用微信「商家转账到零钱」自动打款（T4-6）。
- T4-3 提现回写：微信异步回调收口（成功置 paid）；失败自动换单重试一次，仍失败置 rejected 转人工；
  财务仍可手动确认到账（manual 兜底）或对失败单「补发」。
- T4-4 余额与流水：兼职端本人可读 balances/commissions/withdrawals。
- T4-5 台账/对账导出：按日聚合 commissions + withdrawals，固定 Excel 模板（含佣金列/提现列/汇总行）。

微信自动打款防重铁律（官方文档 2025.03.21）：
遇到 ALREADY_EXISTS / FREQUENCY_LIMIT_* / SYSTEM_ERROR 等错误码必须先查单，
查单明确 FAIL 才允许生成新 out_bill_no 换单重试，否则有重复转账资金风险。
"""
from datetime import datetime, date, timedelta
from decimal import Decimal, InvalidOperation

from flask import Blueprint, current_app, g, request, jsonify
from sqlalchemy.exc import IntegrityError

from ..auth import login_required, require_role
from ..extensions import db
from ..models import Balance, Commission, User, Withdrawal
from .. import wxpay
from ..wxpay import WxpayError
from ..wxpay.transfer import (
    build_transfer_payload,
    classify_create_error,
    normalize_create,
    normalize_query,
)

bp = Blueprint("wallet", __name__)

PAID_SOURCES = ("auto", "manual")


def _to_decimal(value):
    """把输入转为 Decimal，失败抛异常。"""
    if isinstance(value, Decimal):
        return value
    return Decimal(str(value).strip())


def _get_or_create_balance(openid):
    """按 openid 取用户佣金账户，不存在则创建（一人一账户，唯一索引兜底）。"""
    balance = db.session.execute(
        db.select(Balance).where(Balance.openid == openid)
    ).scalar_one_or_none()
    if not balance:
        balance = Balance(openid=openid, available_balance=0)
        db.session.add(balance)
        db.session.flush()
    return balance


# ---------- T4-1 佣金入账（幂等 + 事务，供核验回写复用） ----------

def settle_commission(assignment, task, submitted_quantity):
    """为已通过核验的作业结算佣金，与调用方共享同一事务。

    佣金金额 = 提交数量 × 单价；写 commissions + 余额加，任一失败整体回滚。
    幂等：commissions.assignment_id 唯一索引兜底；已存在则直接返回 None。
    """
    if db.session.execute(
        db.select(Commission).where(Commission.assignment_id == assignment.id)
    ).scalar_one_or_none():
        return None
    amount = _to_decimal(submitted_quantity) * _to_decimal(task.unit_price)
    balance = _get_or_create_balance(assignment.openid)
    commission = Commission(
        openid=assignment.openid,
        task_id=assignment.task_id,
        assignment_id=assignment.id,
        amount=amount,
        settle_date=datetime.utcnow().date(),
    )
    db.session.add(commission)
    balance.available_balance = _to_decimal(balance.available_balance or 0) + amount
    assignment.finish_time = datetime.utcnow()
    return commission


# ---------- T4-4 兼职端：余额与流水查询 ----------

def _commission_view(c):
    return {
        "id": c.id,
        "openid": c.openid,
        "task_id": c.task_id,
        "assignment_id": c.assignment_id,
        "amount": str(c.amount),
        "settle_date": c.settle_date.isoformat() if c.settle_date else None,
        "create_time": c.create_time.isoformat() if c.create_time else None,
    }


def _withdrawal_view(w):
    return {
        "id": w.id,
        "openid": w.openid,
        "amount": str(w.amount),
        "bank_account": w.bank_account,
        "status": w.status,
        "pay_ref_no": w.pay_ref_no,
        "paid_source": w.paid_source,
        "out_bill_no": w.out_bill_no,
        "transfer_bill_no": w.transfer_bill_no,
        "transfer_status": w.transfer_status,
        "transfer_time": w.transfer_time.isoformat() if w.transfer_time else None,
        "fail_reason": w.fail_reason,
        "retry_count": w.retry_count,
        "apply_time": w.apply_time.isoformat() if w.apply_time else None,
        "paid_time": w.paid_time.isoformat() if w.paid_time else None,
    }


@bp.get("/me")
@login_required
def my_balance():
    """兼职端：我的资金概览（余额 + 累计入账 + 累计已提 + 绑卡状态）。"""
    balance = db.session.execute(
        db.select(Balance).where(Balance.openid == g.openid)
    ).scalar_one_or_none()
    available = _to_decimal(balance.available_balance if balance else 0)
    total_income = db.session.execute(
        db.select(db.func.coalesce(db.func.sum(Commission.amount), 0)).where(
            Commission.openid == g.openid
        )
    ).scalar()
    total_withdrawn = db.session.execute(
        db.select(db.func.coalesce(db.func.sum(Withdrawal.amount), 0)).where(
            Withdrawal.openid == g.openid,
            Withdrawal.status.in_(["paid", "pending"]),
        )
    ).scalar()
    return {
        "code": 0,
        "data": {
            "openid": g.openid,
            "available_balance": str(available),
            "total_income": str(total_income),
            "total_withdrawn": str(total_withdrawn),
            "bank_info": g.user.bank_info,
            "has_bank": bool(g.user.bank_info and g.user.bank_info.get("cardNo")),
        },
    }


@bp.post("/bank")
@login_required
def bind_bank():
    """兼职端：绑定/更新收款账户（A5，提现前置条件）。body: {bankName, cardNo, cardHolder}"""
    body = request.get_json(silent=True) or {}
    bank_name = (body.get("bankName") or body.get("bank_name") or "").strip()
    card_no = (body.get("cardNo") or body.get("card_no") or "").strip()
    card_holder = (body.get("cardHolder") or body.get("card_holder") or "").strip()
    if not bank_name or not card_no or not card_holder:
        return {"code": 400, "message": "开户行、卡号、持卡人姓名必填"}, 400
    g.user.bank_info = {"bankName": bank_name, "cardNo": card_no, "cardHolder": card_holder}
    db.session.commit()
    return {"code": 0, "data": g.user.bank_info}


@bp.get("/commissions")
@login_required
def my_commissions():
    """兼职端：本人佣金流水明细。"""
    rows = db.session.execute(
        db.select(Commission)
        .where(Commission.openid == g.openid)
        .order_by(Commission.id.desc())
    ).scalars()
    return {"code": 0, "data": [_commission_view(c) for c in rows]}


# ---------- T4-6 微信自动打款（商家转账到零钱，含防重换单重试） ----------

def _gen_out_bill_no(withdrawal_id, attempt):
    """生成商户转账单号：仅数字+大小写字母，商户内唯一，随 attempt 递增保证换单必为新号。"""
    return f"WD{int(withdrawal_id):010d}A{int(attempt)}"


def _build_payload(client, w, out_bill_no, app):
    """构造发起转账请求体（openid 收款到零钱；金额元→分；姓名加密）。"""
    name_encrypted = None
    user = db.session.execute(
        db.select(User).where(User.openid == w.openid)
    ).scalar_one_or_none()
    real_name = ((user.real_name if user else "") or "").strip() or None
    if real_name:
        name_encrypted = client.encrypt_user_name(real_name)
    return build_transfer_payload(
        appid=app.config.get("WX_APPID", ""),
        out_bill_no=out_bill_no,
        openid=w.openid,
        amount_decimal=w.amount,
        remark=f"兼职佣金提现#{w.id}",
        transfer_scene_id=str(app.config.get("WXPAY_TRANSFER_SCENE_ID", "1000")),
        scene_report_infos=app.config.get("WXPAY_SCENE_REPORT_INFOS")
        or [{"info_type": "提现说明", "info_content": "兼职佣金提现"}],
        notify_url=app.config.get("WXPAY_NOTIFY_URL") or None,
        user_recv_perception="兼职佣金提现",
        user_name_encrypted=name_encrypted,
    )


def _mark_paid(w, out_bill_no, q):
    """转账成功：置 paid，记录商户单号/微信单号/到账时间。"""
    w.status = "paid"
    w.paid_source = "auto"
    w.pay_ref_no = out_bill_no
    w.out_bill_no = out_bill_no
    if q and q.get("transfer_bill_no"):
        w.transfer_bill_no = q["transfer_bill_no"]
    w.transfer_status = "SUCCESS"
    w.fail_reason = None
    w.transfer_time = datetime.utcnow()
    w.paid_time = datetime.utcnow()
    db.session.commit()
    return w.status


def _keep_pending(w, out_bill_no, q, note=None):
    """非终态：仅更新转账字段，status 保持 pending（前端显示「转账中」）。"""
    w.out_bill_no = out_bill_no
    if q and q.get("transfer_bill_no"):
        w.transfer_bill_no = q["transfer_bill_no"]
    w.transfer_status = (q or {}).get("state")
    if note:
        w.fail_reason = note
    db.session.commit()
    return w.status


def _mark_rejected(w, out_bill_no, q, reason=None):
    """最终失败：置 rejected，记录失败原因；余额不退回，由管理员补发/人工处理。"""
    w.status = "rejected"
    w.out_bill_no = out_bill_no
    if q and q.get("transfer_bill_no"):
        w.transfer_bill_no = q["transfer_bill_no"]
    w.transfer_status = "FAIL"
    w.fail_reason = reason or ((q or {}).get("fail_reason")) or "微信转账失败"
    w.transfer_time = datetime.utcnow()
    db.session.commit()
    return w.status


def _persist_transfer(w, out_bill_no, state, bill_no=None, fail_reason=None):
    """落库转账中间状态（用于重试前记录旧单已 FAIL，便于陈旧回调识别）。"""
    w.out_bill_no = out_bill_no
    w.transfer_status = state
    if bill_no:
        w.transfer_bill_no = bill_no
    if fail_reason:
        w.fail_reason = fail_reason
    db.session.commit()


def _do_transfer_attempt(w, out_bill_no):
    """执行单次转账发起 + 错误处理。

    返回 {"state": ..., "q": 归一化结果}：
    - state=SUCCESS       转账成功（终态）
    - state=FAIL          明确失败（查单确认 / 创建即 FAIL）
    - state=REJECTED      确定性错误（换单必败，直接转人工）
    - state=UNKNOWN       结果未知（查单失败/网络异常，保持 pending 待人工）
    - 其余（ACCEPTED/PROCESSING/WAIT_USER_CONFIRM/TRANSFERING/CANCELING）非终态，等回调
    """
    app = current_app
    client = wxpay.get_client(app)
    payload = _build_payload(client, w, out_bill_no, app)
    try:
        resp = client.create_transfer_bill(payload)
        state, bill_no = normalize_create(resp)
        return {"state": state, "q": {"state": state, "transfer_bill_no": bill_no}}
    except WxpayError as exc:
        if classify_create_error(exc.code) != "QUERY_FIRST":
            # 确定性失败，换单也必失败
            return {
                "state": "REJECTED",
                "q": {"state": "FAIL", "fail_reason": f"微信返回[{exc.code}]: {exc.message}"},
            }
        # 结果未知 → 先查单（防重铁律，绝不立即换单）
        try:
            q = normalize_query(client.query_transfer_bill(out_bill_no))
        except WxpayError:
            return {"state": "UNKNOWN", "q": {"state": None, "fail_reason": "查单失败，待人工对账"}}
        return {"state": q["state"], "q": q}


def auto_transfer_withdrawal(w):
    """提现申请后同步自动打款（或 notify 失败后换单重试），任何异常都不向上抛。"""
    try:
        app = current_app
        if wxpay.get_client(app) is None:
            return w.status  # 未启用自动打款 → 保持 pending，沿用 manual 兜底
        retry_max = int(app.config.get("WXPAY_RETRY_MAX", 1))
        while True:
            attempt = int(w.retry_count or 0)
            if attempt > retry_max:
                return w.status
            out_bill_no = _gen_out_bill_no(w.id, attempt)
            result = _do_transfer_attempt(w, out_bill_no)
            state = result["state"]
            if state == "SUCCESS":
                return _mark_paid(w, out_bill_no, result["q"])
            if state == "REJECTED":
                return _mark_rejected(w, out_bill_no, result["q"])
            if state == "UNKNOWN":
                return _keep_pending(w, out_bill_no, result["q"], note=result["q"].get("fail_reason"))
            if state == "FAIL":
                if int(w.retry_count or 0) < retry_max:
                    w.retry_count = int(w.retry_count or 0) + 1
                    _persist_transfer(
                        w, out_bill_no, "FAIL",
                        result["q"].get("transfer_bill_no"),
                        result["q"].get("fail_reason"),
                    )
                    continue  # 查单明确 FAIL → 允许换单重试（新 out_bill_no）
                return _mark_rejected(w, out_bill_no, result["q"])
            # 非终态：ACCEPTED/PROCESSING/WAIT_USER_CONFIRM/TRANSFERING/CANCELING → 等回调
            return _keep_pending(w, out_bill_no, result["q"])
    except Exception:
        db.session.rollback()
        current_app.logger.exception("自动打款异常，提现单 #%s 保持 pending 待人工", w.id)
        return w.status


def _apply_fail(w, out_bill_no, event):
    """notify 回调 FAIL 收口：未达上限自动换单重试一次，否则转人工（余额不退）。"""
    retry_max = int(current_app.config.get("WXPAY_RETRY_MAX", 1))
    reason = event.get("fail_reason") or "微信转账失败"
    if int(w.retry_count or 0) < retry_max:
        w.retry_count = int(w.retry_count or 0) + 1
        _persist_transfer(w, out_bill_no, "FAIL", event.get("transfer_bill_no"), reason)
        auto_transfer_withdrawal(w)  # 换单重试（内部已捕获异常）
        return
    _mark_rejected(
        w, out_bill_no,
        {"state": "FAIL", "transfer_bill_no": event.get("transfer_bill_no")},
        reason,
    )


def handle_transfer_event(out_bill_no, event):
    """notify 回调状态收口（幂等 + 陈旧回调忽略）。返回 True=已处理，False=忽略。"""
    w = db.session.execute(
        db.select(Withdrawal).where(Withdrawal.out_bill_no == out_bill_no)
    ).scalar_one_or_none()
    if w is None:
        current_app.logger.warning("微信回调未找到提现记录，out_bill_no=%s", out_bill_no)
        return False
    if event.get("out_bill_no") != w.out_bill_no:
        current_app.logger.warning(
            "微信陈旧回调已忽略，out_bill_no=%s，当前=%s", out_bill_no, w.out_bill_no
        )
        return False
    state = (event.get("state") or "").upper()
    if state == "SUCCESS":
        if w.status == "paid":
            return True  # 幂等
        _mark_paid(w, out_bill_no, {"state": "SUCCESS", "transfer_bill_no": event.get("transfer_bill_no")})
    elif state == "FAIL":
        if w.status == "rejected":
            return True  # 幂等
        _apply_fail(w, out_bill_no, event)
    else:
        _keep_pending(w, out_bill_no, {"state": state, "transfer_bill_no": event.get("transfer_bill_no")})
    return True


# ---------- T4-2 提现申请（免审核，扣减余额；申请即自动打款） ----------

@bp.post("/withdrawals")
@login_required
def apply_withdrawal():
    """提现申请：校验余额充足且已绑卡，事务内扣余额 + 写 pending 记录（免二次审核）。"""
    body = request.get_json(silent=True) or {}
    amount_raw = body.get("amount")
    try:
        amount = _to_decimal(amount_raw)
    except (TypeError, ValueError, InvalidOperation):
        return {"code": 400, "message": "提现金额无效"}, 400
    if amount <= 0:
        return {"code": 400, "message": "提现金额需为正数"}, 400

    bank = g.user.bank_info
    if not (bank and bank.get("cardNo")):
        return {"code": 400, "message": "请先绑定收款账户"}, 400

    balance = _get_or_create_balance(g.openid)
    if _to_decimal(balance.available_balance or 0) < amount:
        return {"code": 400, "message": "余额不足"}, 400

    # 事务：扣余额 + 写提现记录
    try:
        balance.available_balance = _to_decimal(balance.available_balance) - amount
        record = Withdrawal(
            openid=g.openid,
            amount=amount,
            bank_account=bank,
            status="pending",
            apply_time=datetime.utcnow(),
        )
        db.session.add(record)
        db.session.commit()
    except IntegrityError:
        db.session.rollback()
        return {"code": 500, "message": "提现处理失败，请重试"}, 500

    # 提现申请即全自动打款（HTTP 调用在事务外，失败不阻断申请，转人工兜底）
    auto_transfer_withdrawal(record)
    db.session.refresh(record)
    return {"code": 0, "message": "提现申请成功", "data": _withdrawal_view(record)}


@bp.get("/withdrawals")
@login_required
def my_withdrawals():
    """兼职端：本人提现记录。"""
    rows = db.session.execute(
        db.select(Withdrawal)
        .where(Withdrawal.openid == g.openid)
        .order_by(Withdrawal.id.desc())
    ).scalars()
    return {"code": 0, "data": [_withdrawal_view(w) for w in rows]}


# ---------- T4-3 提现回写（半自动：财务确认到账置 paid） ----------

@bp.post("/withdrawals/<int:withdrawal_id>/confirm")
@require_role("admin")
def confirm_withdrawal(withdrawal_id):
    """财务对 pending 提现执行「确认到账」置为 paid（半自动 manual，D4-2 兜底）。

    body: { pay_ref_no?, paid_source? }（默认 manual）。
    """
    record = db.session.get(Withdrawal, withdrawal_id)
    if not record:
        return {"code": 404, "message": "提现记录不存在"}, 404
    if record.status != "pending":
        return {"code": 400, "message": f"当前状态 {record.status} 不可确认到账"}, 400

    body = request.get_json(silent=True) or {}
    paid_source = (body.get("paid_source") or "manual").strip()
    if paid_source not in PAID_SOURCES:
        return {"code": 400, "message": "paid_source 无效"}, 400

    record.status = "paid"
    record.paid_source = paid_source
    record.pay_ref_no = (body.get("pay_ref_no") or "").strip() or None
    record.paid_time = datetime.utcnow()
    db.session.commit()
    return {"code": 0, "data": _withdrawal_view(record)}


@bp.post("/withdrawals/<int:withdrawal_id>/retry")
@require_role("admin")
def retry_withdrawal(withdrawal_id):
    """失败补发：对 rejected/FAIL（或 pending 无单号的历史残留单）重新发起自动转账。

    若存在旧商户单号，先查单防止重复打款：旧单已 SUCCESS 则直接置 paid。
    """
    record = db.session.get(Withdrawal, withdrawal_id)
    if not record:
        return {"code": 404, "message": "提现记录不存在"}, 404
    if record.status != "rejected" and not (record.status == "pending" and not record.out_bill_no):
        return {"code": 400, "message": f"当前状态 {record.status} 不可补发"}, 400

    # 若有旧单号，先查单：旧单已到账则无需补发（防重复打款）
    if record.out_bill_no:
        client = wxpay.get_client(current_app)
        if client is not None:
            try:
                q = normalize_query(client.query_transfer_bill(record.out_bill_no))
            except WxpayError:
                q = None
            if q and q["state"] == "SUCCESS":
                _mark_paid(record, record.out_bill_no, q)
                return {"code": 0, "message": "原单已到账，无需补发", "data": _withdrawal_view(record)}

    # 重置转账字段后重新发起（不重复扣余额）
    record.status = "pending"
    record.paid_source = "auto"
    record.retry_count = 0
    record.transfer_bill_no = None
    record.transfer_status = None
    record.fail_reason = None
    record.transfer_time = None
    db.session.commit()
    auto_transfer_withdrawal(record)
    db.session.refresh(record)
    return {"code": 0, "data": _withdrawal_view(record)}


@bp.get("/withdrawals/admin")
@require_role("admin")
def admin_withdrawals():
    """管理端：提现记录列表（可按 status 过滤，供提现确认页 T4-3 使用）。"""
    status = (request.args.get("status") or "").strip()
    ok = {"pending", "paid", "rejected"}
    q = db.select(Withdrawal).order_by(Withdrawal.id.desc())
    if status in ok:
        q = q.where(Withdrawal.status == status)
    rows = db.session.execute(q).scalars().all()
    openids = {w.openid for w in rows}
    users = (
        {
            u.openid: u
            for u in db.session.execute(
                db.select(User).where(User.openid.in_(openids))
            ).scalars()
        }
        if openids
        else {}
    )
    data = []
    for w in rows:
        item = _withdrawal_view(w)
        item["user_name"] = users[w.openid].real_name if users.get(w.openid) else None
        data.append(item)
    return {"code": 0, "data": data}


# ---------- T4-5 台账 / 对账导出（按日聚合 commissions + withdrawals） ----------

def _parse_date_range():
    """解析日期范围，默认今天。支持 start_date / end_date（YYYY-MM-DD）。"""
    today = datetime.utcnow().date()
    start_raw = (request.args.get("start_date") or "").strip()
    end_raw = (request.args.get("end_date") or "").strip()
    try:
        start = date.fromisoformat(start_raw) if start_raw else today
        end = date.fromisoformat(end_raw) if end_raw else today
    except ValueError:
        return None, None
    if start > end:
        return None, None
    return start, end


def _aggregate_days(start, end):
    """按日聚合佣金与提现，返回 { days:[...], summary:{} }。"""
    # 佣金按 settle_date，提现按 apply_time 所在日聚合
    comms = db.session.execute(
        db.select(Commission)
        .where(Commission.settle_date >= start, Commission.settle_date <= end)
        .order_by(Commission.settle_date)
    ).scalars().all()
    withdrawals = db.session.execute(
        db.select(Withdrawal).where(
            Withdrawal.apply_time >= datetime.combine(start, datetime.min.time()),
            Withdrawal.apply_time <= datetime.combine(end, datetime.max.time()),
        )
    ).scalars().all()

    comm_by_day = {}
    for c in comms:
        d = c.settle_date
        comm_by_day.setdefault(d, {"count": 0, "amount": Decimal("0")})
        comm_by_day[d]["count"] += 1
        comm_by_day[d]["amount"] += _to_decimal(c.amount)

    wd_by_day = {}
    for w in withdrawals:
        d = w.apply_time.date()
        wd_by_day.setdefault(d, {"count": 0, "amount": Decimal("0")})
        wd_by_day[d]["count"] += 1
        wd_by_day[d]["amount"] += _to_decimal(w.amount)

    days = []
    cur = start
    while cur <= end:
        c = comm_by_day.get(cur, {"count": 0, "amount": Decimal("0")})
        w = wd_by_day.get(cur, {"count": 0, "amount": Decimal("0")})
        net = c["amount"] - w["amount"]
        days.append(
            {
                "date": cur.isoformat(),
                "commission_count": c["count"],
                "commission_amount": str(c["amount"]),
                "withdrawal_count": w["count"],
                "withdrawal_amount": str(w["amount"]),
                "net_amount": str(net),
            }
        )
        cur += timedelta(days=1)

    total_comm = sum((Decimal(d["commission_amount"]) for d in days), Decimal("0"))
    total_wd = sum((Decimal(d["withdrawal_amount"]) for d in days), Decimal("0"))
    total_wd_count = sum(d["withdrawal_count"] for d in days)
    summary = {
        "commission_count": sum(d["commission_count"] for d in days),
        "commission_amount": str(total_comm),
        "withdrawal_count": total_wd_count,
        "withdrawal_amount": str(total_wd),
        "net_amount": str(total_comm - total_wd),
    }
    return {"days": days, "summary": summary}


def _build_excel(data):
    """用 openpyxl 生成台账 Excel 报表。未安装则返回 None（调用方回退 CSV）。"""
    try:
        from openpyxl import Workbook
    except ImportError:
        return None
    wb = Workbook()
    ws = wb.active
    ws.title = "对账汇总"
    ws.append(["日期", "佣金入账笔数", "佣金金额", "提现笔数", "提现金额", "净结余"])
    for d in data["days"]:
        ws.append(
            [
                d["date"],
                d["commission_count"],
                float(d["commission_amount"]),
                d["withdrawal_count"],
                float(d["withdrawal_amount"]),
                float(d["net_amount"]),
            ]
        )
    s = data["summary"]
    ws.append(["汇总", s["commission_count"], float(s["commission_amount"]),
               s["withdrawal_count"], float(s["withdrawal_amount"]), float(s["net_amount"])])

    # 明细页（追溯）：佣金
    ws2 = wb.create_sheet("佣金流水")
    ws2.append(["入账日期", "openid", "任务ID", "作业ID", "金额"])
    for c in db.session.execute(
        db.select(Commission).order_by(Commission.id).scalars()
    ):
        ws2.append([c.settle_date.isoformat() if c.settle_date else None,
                    c.openid, c.task_id, c.assignment_id, float(c.amount)])
    # 明细页：提现
    ws3 = wb.create_sheet("提现流水")
    ws3.append(["申请时间", "openid", "金额", "状态", "到账时间"])
    for w in db.session.execute(
        db.select(Withdrawal).order_by(Withdrawal.id).scalars()
    ):
        ws3.append([w.apply_time.isoformat() if w.apply_time else None,
                    w.openid, float(w.amount), w.status,
                    w.paid_time.isoformat() if w.paid_time else None])

    import io
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


@bp.get("/ledger")
@require_role("admin")
def ledger():
    """台账/对账（T4-5，对接 T5-5）。

    默认返回按日聚合的 JSON；format=excel|csv 导出文件。
    固定模板：包含「佣金列 / 提现列 / 汇总行」，可追溯。
    """
    start, end = _parse_date_range()
    if not start:
        return {"code": 400, "message": "日期范围无效"}, 400
    data = _aggregate_days(start, end)

    fmt = (request.args.get("format") or "json").strip().lower()
    if fmt == "excel":
        buf = _build_excel(data)
        name = f"对账台账_{start}_{end}.xlsx"
        from flask import send_file
        if buf is not None:
            return send_file(
                buf,
                as_attachment=True,
                download_name=name,
                mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        # 未装 openpyxl：回退 CSV
        fmt = "csv"
    if fmt == "csv":
        import io
        buf = io.StringIO()
        buf.write("日期,佣金笔数,佣金金额,提现笔数,提现金额,净结余\n")
        for d in data["days"]:
            buf.write(
                f'{d["date"]},{d["commission_count"]},{d["commission_amount"]},'
                f'{d["withdrawal_count"]},{d["withdrawal_amount"]},{d["net_amount"]}\n'
            )
        s = data["summary"]
        buf.write(
            f'汇总,{s["commission_count"]},{s["commission_amount"]},'
            f'{s["withdrawal_count"]},{s["withdrawal_amount"]},{s["net_amount"]}\n'
        )
        from flask import Response
        buf.seek(0)
        return Response(
            buf.getvalue(),
            mimetype="text/csv; charset=utf-8",
            headers={"Content-Disposition": f'attachment; filename="ledger_{start}_{end}.csv"'},
        )
    return jsonify({"code": 0, "data": data})